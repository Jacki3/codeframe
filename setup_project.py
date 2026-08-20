"""Step 1: work out what the raw data contains, and split it into a project.

    python setup_project.py --raw "../"  --to ../myproject            # inspect
    python setup_project.py --to ../myproject --apply                 # build

Two phases, deliberately separate.

INSPECT reads whatever is in the raw folder - spreadsheets, CSVs, transcripts -
and writes setup.json: everything it found, plus a proposed mapping saying which
column is the participant id, which identifies the unit each person did, which
columns hold the text to code, which are categories to compare across, and which
are numeric measures. Nothing else is written and nothing is guessed at silently:
every proposal carries the evidence it was made from.

You then read setup.json and correct it. It is a small file and it is meant to be
edited - the machine is good at finding candidates and bad at knowing which of
two plausible columns you actually meant.

APPLY reads the corrected setup.json and writes sources.csv and frame.csv.

WHY THE MAPPING HAS THESE FIVE ROLES
    pid         who said it
    unit        what they did that this response is about - a game, a condition,
                a session. Together (pid, unit) is one row of the frame, and the
                denominator for every rate the findings will report.
    text        columns to be read and quoted from. Each becomes a source document.
    categories  things to split findings by later: device, gender, group.
    measures    numeric scores to compare codes against: scales, subscales, timings.

A study with no unit - one response per person - still works: leave unit empty and
the frame is one row per participant.
"""
import argparse, collections, csv, json, os, re, statistics, sys

# Quotes carry curly apostrophes and dashes; a cp1252 console mangles them on the
# way to the screen even though the files are fine.
try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except (AttributeError, ValueError):
    pass

HERE = os.path.dirname(os.path.abspath(__file__))

TEXTUAL = 40          # mean characters above which a column is prose, not a label
MAX_CATEGORY = 12     # distinct values above which a column stops being a category


# ---------------------------------------------------------------- reading

def read_tabular(path):
    """Return {sheet_name: (headers, rows)} for a csv or xlsx."""
    ext = os.path.splitext(path)[1].lower()
    if ext in (".csv", ".tsv", ".txt"):
        raw = open(path, "rb").read()
        for enc in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
            try:
                text = raw.decode(enc); break
            except UnicodeDecodeError:
                continue
        else:
            return {}
        import io
        rows = list(csv.reader(io.StringIO(text, newline="")))
        if not rows:
            return {}
        return {os.path.basename(path): (rows[0], rows[1:])}
    if ext in (".xlsx", ".xlsm"):
        try:
            import openpyxl
        except ImportError:
            print("  (openpyxl not installed - skipping xlsx)", file=sys.stderr)
            return {}
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        out = {}
        for name in wb.sheetnames:
            rows = [list(r) for r in wb[name].iter_rows(values_only=True)]
            rows = [r for r in rows if any(c not in (None, "") for c in r)]
            if len(rows) > 1:
                out[name] = ([str(c) if c is not None else "" for c in rows[0]], rows[1:])
        return out
    return {}


def profile(headers, rows):
    """Describe each column: how full, how varied, how long, how numeric."""
    cols = []
    for i, h in enumerate(headers):
        vals = [r[i] for r in rows if i < len(r)]
        vals = [v for v in vals if v not in (None, "")]
        if not vals:
            cols.append({"index": i, "header": h, "filled": 0, "distinct": 0,
                         "mean_len": 0, "numeric": False, "examples": []})
            continue
        strs = [str(v).strip() for v in vals]
        numeric = sum(1 for v in strs if re.fullmatch(r"-?\d+(\.\d+)?", v)) / len(strs) > 0.9
        cols.append({
            "index": i, "header": h, "filled": len(vals),
            "distinct": len(set(strs)),
            "mean_len": round(statistics.mean(len(s) for s in strs), 1),
            "numeric": numeric,
            "examples": [s[:60] for s in list(dict.fromkeys(strs))[:4]],
        })
    return cols


# ---------------------------------------------------------------- proposing

PID_HINT = re.compile(r"participant|\bpid\b|\bid\b|respondent", re.I)
UNIT_HINT = re.compile(r"game|title|condition|task|app|session|version", re.I)
SKIP_HINT = re.compile(r"^(start|completion|submitted|timestamp)", re.I)


def propose(cols, n_rows):
    """Guess each column's role, and say why. Heuristics, not certainty."""
    p = {"pid": None, "unit": None, "text": [], "categories": [], "measures": [],
         "ignored": [], "why": {}}

    # A participant id must REPEAT - the same person answering about two games
    # appears twice - and must be more granular than a category. A per-row key
    # like "Response ID" has one value per row, which is what disqualifies it.
    def pid_rank(c):
        h = c["header"].lower()
        # an explicit "PID" beats "participant", which beats a bare "...Id"
        return (2 if "pid" in re.findall("[a-z]+", h) else
                1 if "participant" in h else 0, c["distinct"])

    pid_cands = [c for c in cols
                 if c["filled"] and PID_HINT.search(c["header"])
                 and MAX_CATEGORY < c["distinct"] <= n_rows * 0.9]
    if pid_cands:
        best = max(pid_cands, key=pid_rank)
        p["pid"] = best["header"]
        per = n_rows / best["distinct"]
        p["why"][best["header"]] = (
            f"repeats: {best['distinct']} distinct over {n_rows} rows "
            f"(~{per:.1f} responses each), and named like a participant id")

    for c in cols:
        h, why = c["header"], None
        if h == p["pid"]:
            continue
        if not c["filled"] or SKIP_HINT.search(h):
            p["ignored"].append(h); continue

        if p["unit"] is None and UNIT_HINT.search(h) and 1 < c["distinct"] <= MAX_CATEGORY:
            p["unit"] = h
            why = f"named like a unit and takes {c['distinct']} values: {', '.join(c['examples'][:4])}"
        elif c["mean_len"] >= TEXTUAL and c["distinct"] > n_rows * 0.5:
            p["text"].append(h)
            why = f"prose: {c['mean_len']} chars on average, mostly unique"
        elif c["numeric"]:
            p["measures"].append(h)
            why = "numeric"
        elif 1 < c["distinct"] <= MAX_CATEGORY:
            p["categories"].append(h)
            why = f"{c['distinct']} values: {', '.join(c['examples'][:4])}"
        else:
            p["ignored"].append(h)
            why = f"{c['distinct']} distinct values, mean length {c['mean_len']}"
        p["why"][h] = why
    return p


NAME_HINT = re.compile(r"name|email|e-mail|student|address", re.I)


def find_roster(tables):
    """Find the table that links a participant id to something matchable.

    The response table is usually anonymous, so the route from a recording to a
    PID lives in a separate roster. Two rankings matter and both bit me:

    - within a table, an explicit "PID" column must beat an internal key like
      "EOI Id". They are different numbering schemes, and picking the wrong one
      mislabels every quote that gets linked.
    - across tables, a roster keyed on a real PID beats one that is not.
    """
    def id_rank(header):
        h = header.lower()
        return 2 if "pid" in re.findall("[a-z]+", h) else 1 if "participant" in h else 0

    cands = []
    for t in tables:
        cols = {c["header"]: c for c in t["columns"]}
        idish = [h for h, c in cols.items()
                 if PID_HINT.search(h) and c["distinct"] > MAX_CATEGORY]
        match_cols = [h for h, c in cols.items()
                      if NAME_HINT.search(h) and c["mean_len"] > 4
                      and c["distinct"] > MAX_CATEGORY]
        if not idish or not match_cols:
            continue
        pid = max(idish, key=lambda h: (id_rank(h), cols[h]["distinct"]))
        score = (id_rank(pid), len(match_cols))
        cands.append((score, {"file": t["file"], "sheet": t["sheet"],
                              "pid": pid, "match_columns": match_cols}))
    cands.sort(key=lambda x: x[0], reverse=True)
    return [c[1] for c in cands]


def link_transcripts(transcripts, roster, raw):
    """Propose which participant each transcript belongs to, and show the evidence.

    Two routes, strongest first: a student number appearing in the file and in the
    roster, or name words shared between the roster and the file path. Anything
    unmatched is listed rather than guessed at.
    """
    if not roster:
        return [{"path": t["path"], "pid": "", "via": "no roster found",
                 "evidence": ""} for t in transcripts]
    tables = read_tabular(os.path.join(raw, roster["file"]))
    headers, rows = tables.get(roster["sheet"], next(iter(tables.values())))
    ix = {h: i for i, h in enumerate(headers)}
    people = []
    for row in rows:
        def cell(h):
            i = ix.get(h)
            return "" if i is None or i >= len(row) or row[i] is None else str(row[i]).strip()
        pid = cell(roster["pid"])
        if not pid:
            continue
        blob = " ".join(cell(h) for h in roster["match_columns"])
        people.append({"pid": re.sub(r"\D", "", pid) or pid, "blob": blob,
                       "ids": set(re.findall("[sS][0-9]{6,8}", blob)),
                       "words": {w.lower() for w in re.findall(r"[A-Za-z]{4,}", blob)}})

    links = []
    for t in transcripts:
        hay = f"{t['path']} {t['folder']} {' '.join(t['speakers'])}"
        tids = {i.lower() for i in t["student_ids"]}
        hit, via, ev = None, "", ""
        for p in people:
            if tids & {i.lower() for i in p["ids"]}:
                hit, via = p, "student number"
                ev = ", ".join(sorted(tids & {i.lower() for i in p["ids"]}))
                break
        if not hit:
            hay_words = {w.lower() for w in re.findall(r"[A-Za-z]{4,}", hay)}
            scored = [(len(p["words"] & hay_words), p) for p in people]
            scored = [x for x in scored if x[0] >= 2]
            if len(scored) == 1 or (scored and sorted(scored, reverse=True,
                                                      key=lambda x: x[0])[0][0] >= 2):
                scored.sort(key=lambda x: -x[0])
                if len(scored) == 1 or scored[0][0] > scored[1][0]:
                    hit, via = scored[0][1], "name match"
                    ev = ", ".join(sorted(scored[0][1]["words"] & hay_words))
        links.append({"path": t["path"], "pid": hit["pid"] if hit else "",
                      "via": via or "no match - fill in the pid by hand",
                      "evidence": ev})
    return links


TIMECODE = re.compile(r"\d\d:\d\d:\d\d[.,]\d+\s*-->|^\[\d+:\d\d\]", re.M)


def find_transcripts(root, dst=None, skip=()):
    """Transcript files, and whoever appears to be speaking in them.

    A .txt only counts if it actually contains timecodes or speaker markers -
    otherwise every notes file and coding batch in the tree looks like a
    transcript, which is exactly what happened the first time.
    """
    out = []
    for dirpath, dirs, files in os.walk(root):
        dirs[:] = [d for d in dirs
                   if not d.startswith(".")
                   and d != "__pycache__"
                   and os.path.abspath(os.path.join(dirpath, d)) != HERE
                   and d not in skip
                   and (dst is None or os.path.abspath(os.path.join(dirpath, d))
                        != os.path.abspath(dst))
                   and not os.path.exists(os.path.join(dirpath, d, "sources.csv"))]
        for fn in files:
            if not fn.lower().endswith((".vtt", ".txt", ".srt")):
                continue
            path = os.path.join(dirpath, fn)
            try:
                text = open(path, encoding="utf-8", errors="replace").read()
            except OSError:
                continue
            if fn.lower().endswith(".txt") and not TIMECODE.search(text):
                continue
            speakers = sorted(set(re.findall(r"<v ([^>]+)>", text)))
            ids = sorted(set(re.findall(r"\bs\d{6,8}\b", text, re.I)))
            out.append({
                "path": os.path.relpath(path, root).replace("\\", "/"),
                "folder": os.path.basename(dirpath),
                "speakers": speakers, "student_ids": ids,
                "words": len(re.findall(r"<v [^>]+>(.*?)</v>", text, re.S)),
            })
    return out


CUE = re.compile(r"(\d\d):(\d\d):(\d\d)[.,]\d+ --> .*?<v ([^>]+)>(.*?)</v>", re.S)


def read_vtt(path, pid, roster_blob):
    """Turns from a VTT, with speaker names replaced by roles.

    Participant names never enter sources.csv. Whoever the transcript links to
    becomes PID<n>; everyone else becomes Interviewer. That keeps the corpus
    quotable without carrying identities around with it.
    """
    text = open(path, encoding="utf-8", errors="replace").read()
    words = {w.lower() for w in re.findall("[A-Za-z]{4,}", roster_blob)}
    ids = {i.lower() for i in re.findall("[sS][0-9]{6,8}", roster_blob)}
    out = []
    for hh, mm, ss, who, body in CUE.findall(text):
        body = " ".join(body.split())
        if not body:
            continue
        w = {x.lower() for x in re.findall("[A-Za-z]{4,}", who)}
        i = {x.lower() for x in re.findall("[sS][0-9]{6,8}", who)}
        is_participant = bool(i & ids) or bool(w & words)
        t = int(hh) * 3600 + int(mm) * 60 + int(ss)
        out.append((t, f"PID{pid}" if is_participant else "Interviewer", body))
    # A name spoken aloud survives speaker-label replacement, so remove it from
    # the words too. Only names four characters or longer, to avoid mangling
    # ordinary words that happen to appear in a roster field.
    for w in sorted({x for x in re.findall("[A-Za-z]{4,}", roster_blob)}, key=len,
                    reverse=True):
        if w.lower() in ("please", "provide", "your", "university", "email",
                         "address", "full", "name", "travel", "arrangement",
                         "purposes", "contact", "regarding", "plans", "study"):
            continue
        out = [(t, who, re.sub(w, f"PID{pid}", body, flags=re.I))
               for t, who, body in out]

    merged = []
    for t, who, body in out:
        if merged and merged[-1][1] == who:
            merged[-1][2] += " " + body
        else:
            merged.append([t, who, body])
    return merged


# ---------------------------------------------------------------- phases

def inspect(raw, dst, skip=()):
    print(f"scanning {raw}\n")
    tables = []
    for dirpath, dirs, files in os.walk(raw):
        # skip our own output and anything that is already a project
        dirs[:] = [d for d in dirs
                   if not d.startswith(".")
                   and d != "__pycache__"
                   and os.path.abspath(os.path.join(dirpath, d)) != HERE
                   and d not in skip
                   and os.path.abspath(os.path.join(dirpath, d)) != os.path.abspath(dst)
                   and not os.path.exists(os.path.join(dirpath, d, "sources.csv"))]
        for fn in sorted(files):
            if fn.startswith("~$") or not fn.lower().endswith(
                    (".csv", ".tsv", ".xlsx", ".xlsm")):
                continue
            path = os.path.join(dirpath, fn)
            for sheet, (headers, rows) in read_tabular(path).items():
                cols = profile(headers, rows)
                tables.append({
                    "file": os.path.relpath(path, raw).replace("\\", "/"),
                    "sheet": sheet, "rows": len(rows), "columns": cols,
                    "proposed": propose(cols, len(rows)),
                })
                print(f"  {os.path.relpath(path, raw)} [{sheet}]  "
                      f"{len(rows)} rows x {len(headers)} columns")

    transcripts = find_transcripts(raw, dst, skip)
    if transcripts:
        print(f"  {len(transcripts)} transcript file(s)")

    # the widest table with both a pid and prose is the likely response table
    ranked = sorted(tables, key=lambda t: (bool(t["proposed"]["pid"]),
                                           len(t["proposed"]["text"]), t["rows"]),
                    reverse=True)
    primary = ranked[0] if ranked else None

    rosters = find_roster(tables)
    roster = rosters[0] if rosters else None
    links = link_transcripts(transcripts, roster, raw) if transcripts else []
    for l in links:
        # Never trusted. A wrong pid mislabels every quote from that transcript,
        # so the researcher confirms the whole block before anything is imported.
        l["confirmed"] = False
    spec = {
        "raw": os.path.abspath(raw),
        "responses": None,
        "roster": roster,
        "roster_candidates": rosters,
        "transcripts_confirmed": False,
        "transcripts": {"kind": "interview", "speaker_prefix": True, "links": links},
        "found": tables,
    }
    if primary:
        pp = primary["proposed"]
        spec["responses"] = {
            "file": primary["file"], "sheet": primary["sheet"],
            "pid": pp["pid"], "unit": pp["unit"], "text": pp["text"],
            "categories": pp["categories"], "measures": pp["measures"],
            "kind": "survey",
        }

    os.makedirs(dst, exist_ok=True)
    out = os.path.join(dst, "setup.json")
    with open(out, "w", encoding="utf-8") as f:
        json.dump(spec, f, indent=1)

    print(f"\nproposed mapping (from {primary['file']} [{primary['sheet']}])"
          if primary else "\nno table with both an id and prose was found")
    if primary:
        pp = primary["proposed"]
        for role in ("pid", "unit"):
            v = pp[role]
            print(f"  {role:<11} {v or '(none found)'}")
            if v:
                print(f"              why: {pp['why'].get(v, '')}")
        for role in ("text", "categories", "measures"):
            names = pp[role]
            print(f"  {role:<11} {len(names)}")
            for h in names[:4]:
                print(f"              {h[:64]}")
            if len(names) > 4:
                print(f"              ... and {len(names) - 4} more")
    print(f"\nwrote {out}")
    print("Read it, correct anything wrong, then re-run with --apply.")
    print("The machine finds candidates; only you know which column you meant.")


def apply(dst):
    spec_path = os.path.join(dst, "setup.json")
    if not os.path.exists(spec_path):
        raise SystemExit(f"no setup.json in {dst} - run the inspect phase first")
    spec = json.load(open(spec_path, encoding="utf-8"))
    r = spec.get("responses")
    if not r or not r.get("pid"):
        raise SystemExit("setup.json has no responses.pid - fill it in and re-run")

    path = os.path.join(spec["raw"], r["file"])
    tables = read_tabular(path)
    headers, rows = tables[r["sheet"]] if r["sheet"] in tables else next(iter(tables.values()))
    ix = {h: i for i, h in enumerate(headers)}

    def cell(row, header):
        i = ix.get(header)
        v = row[i] if i is not None and i < len(row) else None
        return "" if v is None else str(v).strip()

    # A survey export writes the whole question into the header, and that string
    # would otherwise become a frame column and a label on every finding. The
    # review step proposes short names; without it the header is used as-is.
    lab = r.get("labels") or {}

    def fname(h):
        return lab.get(h, h)

    sources, frame = [], {}
    for row in rows:
        pid = cell(row, r["pid"])
        if not pid:
            continue
        unit = cell(row, r["unit"]) if r.get("unit") else ""
        key = (pid, unit)
        frame.setdefault(key, {
            "pid": pid, "unit": unit,
            **{fname(h): cell(row, h) for h in r.get("categories", [])},
            **{fname(h): cell(row, h) for h in r.get("measures", [])},
        })
        for h in r.get("text", []):
            body = cell(row, h)
            if not body:
                continue
            label = re.sub(r"^\d+[\.\)]\s*", "", h)[:60]
            sources.append({
                "source_id": f"S-{int(pid):03d}-{len(sources):04d}" if pid.isdigit()
                else f"S-{pid}-{len(sources):04d}",
                "pid": pid, "unit": unit, "kind": r.get("kind", "response"),
                "label": label, "text": body,
            })

    # ---- transcripts, once their pids have been confirmed by hand ----
    kind = spec.get("transcripts", {}).get("kind", "interview")
    links = spec.get("transcripts", {}).get("links") or []
    n_tr = 0
    if links and spec.get("transcripts_confirmed"):
        blobs = {}
        ros = spec.get("roster")
        if ros:
            rt = read_tabular(os.path.join(spec["raw"], ros["file"]))
            rh, rr = rt.get(ros["sheet"], next(iter(rt.values())))
            rix = {h: i for i, h in enumerate(rh)}
            for row in rr:
                def rc(h):
                    i = rix.get(h)
                    return "" if i is None or i >= len(row) or row[i] is None else str(row[i])
                key = re.sub(r"[^0-9]", "", rc(ros["pid"]))
                if key:
                    blobs[key] = " ".join(rc(h) for h in ros["match_columns"])
        for l in links:
            pid = str(l.get("pid") or "").strip()
            if not pid:
                continue
            path = os.path.join(spec["raw"], l["path"])
            if not os.path.exists(path):
                continue
            turns = read_vtt(path, pid, blobs.get(pid, ""))
            if not turns:
                continue
            gap = chr(10) * 2
            body = gap.join(f"[{t // 60}:{t % 60:02d}] {who}: {txt}"
                            for t, who, txt in turns)
            sources.append({
                "source_id": f"I-{int(pid):03d}" if pid.isdigit() else f"I-{pid}",
                "pid": pid, "unit": "", "kind": kind,
                "label": f"{len(turns)} turns, {turns[-1][0] // 60} min", "text": body,
            })
            n_tr += 1
            for (fp, fu), fr in frame.items():
                if fp == pid:
                    fr["has_" + kind] = "yes"
        for fr in frame.values():
            fr.setdefault("has_" + kind, "no")

    def write(p, fields, rowset):
        with open(p, "w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=fields); w.writeheader()
            w.writerows([{k: x.get(k, "") for k in fields} for x in rowset])

    os.makedirs(os.path.join(dst, "data"), exist_ok=True)
    os.makedirs(os.path.join(dst, "codebook"), exist_ok=True)
    write(os.path.join(dst, "sources.csv"),
          ["source_id", "pid", "unit", "kind", "label", "text"], sources)
    frame_fields = (["pid", "unit"]
                    + ([f"has_{kind}"] if n_tr else [])
                    + [fname(h) for h in r.get("categories", [])]
                    + [fname(h) for h in r.get("measures", [])])
    write(os.path.join(dst, "frame.csv"), frame_fields,
          sorted(frame.values(), key=lambda x: (x["pid"], x["unit"])))

    # Which frame columns are categories and which are measures, recorded rather
    # than left to be guessed later. Nothing downstream can recover it from
    # frame.csv alone: play order and a 0-4 rating are both small integers, and
    # the difference between them is a fact about the study, decided here.
    meta = {"categories": ([f"has_{kind}"] if n_tr else [])
                          + [fname(h) for h in r.get("categories", [])],
            "measures": [fname(h) for h in r.get("measures", [])]}
    meta["facets"] = meta["categories"]
    # What this study calls its unit, and what kinds of source it holds. Without
    # these every tool downstream has to guess, and a tool that guesses says
    # "game-play" to someone studying museum visits. The label was already
    # proposed for the unit column during review; it was just never used.
    meta["unit_label"] = (lab.get(r["unit"]) if r.get("unit") else "") or "unit"
    meta["kinds"] = sorted({s_["kind"] for s_ in sources if s_.get("kind")})
    with open(os.path.join(dst, "frame.json"), "w", encoding="utf-8") as f:
        json.dump(meta, f, indent=1)

    print(f"wrote {dst}")
    if n_tr:
        print(f"  imported {n_tr} transcripts as kind='{kind}'")
    print(f"  sources.csv  {len(sources)} documents, "
          f"{sum(len(s['text'].split()) for s in sources):,} words")
    print(f"  frame.csv    {len(frame)} rows ({r['pid']} x {r.get('unit') or 'one per participant'})")
    for h in r.get("categories", [])[:8]:
        vals = collections.Counter(f.get(fname(h), "") for f in frame.values()
                                   if f.get(fname(h)))
        print(f"    {fname(h)[:40]:<42} {len(vals)} values")
    if r.get("measures"):
        print(f"    {len(r['measures'])} numeric measures per row")
    if spec.get("transcripts", {}).get("links") and not spec.get("transcripts_confirmed"):
        n = len(spec["transcripts"]["links"])
        print(f"\n  {n} transcripts NOT imported: transcripts_confirmed is false.")
        print("  Check every pid in setup.json, then set it to true and re-run.")
    unlinked = [l for l in links if not str(l.get("pid") or "").strip()]
    if unlinked:
        print(f"\n  {len(unlinked)} transcript(s) had no pid and were skipped:")
        for l in unlinked[:5]:
            print(f"    {l['path']}")
        print("  Fill in the pid in setup.json and re-run to include them.")


def main():
    ap = argparse.ArgumentParser(description="Inspect raw data and set up a project.")
    ap.add_argument("--raw", help="folder holding the raw data (inspect phase)")
    ap.add_argument("--to", required=True, help="project directory to create")
    ap.add_argument("--skip", default="", metavar="DIRS",
                    help="comma-separated folder names to leave out of the scan")
    ap.add_argument("--review", action="store_true",
                    help="ask Claude to check the proposed mapping")
    ap.add_argument("--dry-run", action="store_true",
                    help="with --review: print what would be sent and send nothing")
    ap.add_argument("--model", default="sonnet", help="model for --review")
    ap.add_argument("--apply", action="store_true", help="build from a corrected setup.json")
    a = ap.parse_args()
    if a.review:
        import review
        review.review(a.to, model=a.model, dry_run=a.dry_run)
    elif a.apply:
        apply(a.to)
    elif a.raw:
        inspect(a.raw, a.to, tuple(x.strip() for x in a.skip.split(',') if x.strip()))
    else:
        ap.error("give --raw to inspect, --review to check it, or --apply to build")


if __name__ == "__main__":
    main()
